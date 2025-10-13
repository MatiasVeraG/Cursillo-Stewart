import pandas as pd
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class ExcelTemplateGeneratorV3:
    """Generador de plantillas Excel v3 con formato específico"""
    
    def __init__(self):
        self.template_path = 'documents/Plantilla_de_Ingresantes_v3.xlsx'
        
    def generate_template(self, exam_name="", year="", with_sample_data=False):
        """
        Genera plantilla Excel v3 con formato específico
        - A1:E1 merge: "Lista de Ingresantes" (centrado/bold)
        - Fila 2: cabeceras exactas
        - G2/H2: Nombre de Examen / valor
        - G3/H3: Año / valor
        - Datos desde fila 3
        """
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Ingresantes"
        
        # A1:E1 - Título principal (merge y centrado)
        ws.merge_cells('A1:E1')
        ws['A1'] = "Lista de Ingresantes"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Fila 2 - Cabeceras exactas
        headers = ["Nombre y Apellido", "Puntaje", "Carrera", "Puesto", "Preferencial"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=i, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        # Metadatos en columnas G y H
        # G2/H2: Nombre de Examen
        ws['G2'] = "Nombre de Examen"
        ws['G2'].font = Font(bold=True)
        ws['H2'] = exam_name if exam_name else "UPTP"
        
        # G3/H3: Año
        ws['G3'] = "Año"
        ws['G3'].font = Font(bold=True)
        ws['H3'] = year if year else "2025"
        
        # Datos de ejemplo (si se solicita)
        if with_sample_data:
            sample_data = [
                ["Juan Pérez García", 95.5, "Ing. Informática", 1, "Sí"],
                ["María González de la Cruz", 92.8, "Ing. Civil", 2, "No"],
                ["Carlos López", 89.3, "Ing. Electromecánica", 3, "Sí"],
                ["Ana Rodríguez", 87.1, "Ing. Industrial", 4, "No"],
                ["Luis Fernández del Valle", 85.9, "Ing. Informática", 5, "Sí"]
            ]
            
            for i, row_data in enumerate(sample_data, 3):  # Empezar en fila 3
                for j, value in enumerate(row_data, 1):
                    ws.cell(row=i, column=j, value=value)
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 25,  # Nombre y Apellido
            'B': 12,  # Puntaje
            'C': 20,  # Carrera
            'D': 10,  # Puesto
            'E': 15,  # Preferencial
            'G': 18,  # Labels de metadatos
            'H': 15   # Valores de metadatos
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Guardar archivo
        os.makedirs(os.path.dirname(self.template_path), exist_ok=True)
        wb.save(self.template_path)
        
        return {
            'success': True,
            'path': self.template_path,
            'exam': exam_name or "UPTP",
            'year': year or "2025",
            'has_sample_data': with_sample_data
        }
    
    def read_excel_v3(self, file_path):
        """
        Lee un archivo Excel v3 y extrae metadatos y datos
        """
        try:
            # Leer metadatos desde celdas específicas
            df_meta = pd.read_excel(file_path, header=None)
            
            # Extraer metadatos de G2/H2 y G3/H3
            exam_name = ""
            year = ""
            
            try:
                if len(df_meta) > 1 and len(df_meta.columns) > 7:
                    exam_name = str(df_meta.iloc[1, 7]) if pd.notna(df_meta.iloc[1, 7]) else ""
                if len(df_meta) > 2 and len(df_meta.columns) > 7:
                    year = str(df_meta.iloc[2, 7]) if pd.notna(df_meta.iloc[2, 7]) else ""
            except:
                pass
            
            # Leer datos principales desde fila 2 (header=1)
            df_data = pd.read_excel(file_path, header=1)
            
            # Filtrar filas completamente vacías
            df_clean = df_data.dropna(how='all')
            
            # Convertir a lista de diccionarios
            records = []
            for _, row in df_clean.iterrows():
                if pd.notna(row.get('Nombre y Apellido', '')):
                    record = {
                        'Nombre y Apellido': row.get('Nombre y Apellido', ''),
                        'Puntaje': row.get('Puntaje', 0),
                        'Carrera': row.get('Carrera', ''),
                        'Puesto': row.get('Puesto', ''),
                        'Preferencial': row.get('Preferencial', '')
                    }
                    records.append(record)
            
            return {
                'success': True,
                'metadata': {
                    'exam': exam_name,
                    'year': year
                },
                'records': records,
                'total_records': len(records)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'metadata': {'exam': '', 'year': ''},
                'records': [],
                'total_records': 0
            }

def main():
    generator = ExcelTemplateGeneratorV3()
    
    print("=== GENERADOR DE PLANTILLA EXCEL V3 ===")
    print()
    
    # Generar plantilla básica
    print("1. Generando plantilla básica...")
    result = generator.generate_template()
    
    if result['success']:
        print(f"✅ Plantilla generada: {result['path']}")
        print(f"📋 Examen: {result['exam']}")
        print(f"📅 Año: {result['year']}")
        print()
        
        # Generar plantilla con datos de ejemplo
        print("2. Generando plantilla con datos de ejemplo...")
        sample_path = 'documents/Plantilla_con_Ejemplos_v3.xlsx'
        generator.template_path = sample_path
        result_sample = generator.generate_template("MOFA", "2026", with_sample_data=True)
        
        if result_sample['success']:
            print(f"✅ Plantilla con ejemplos: {result_sample['path']}")
            print()
            
            # Probar lectura
            print("3. Probando lectura de plantilla con ejemplos...")
            read_result = generator.read_excel_v3(sample_path)
            
            if read_result['success']:
                print(f"📋 Examen detectado: {read_result['metadata']['exam']}")
                print(f"📅 Año detectado: {read_result['metadata']['year']}")
                print(f"👥 Registros encontrados: {read_result['total_records']}")
                print()
                print("Primeros registros:")
                for i, record in enumerate(read_result['records'][:3]):
                    print(f"  {i+1}. {record['Nombre y Apellido']} - {record['Puntaje']} - {record['Carrera']}")
            else:
                print(f"❌ Error leyendo: {read_result['error']}")
    else:
        print("❌ Error generando plantilla")

if __name__ == "__main__":
    main()